import datetime
from django.shortcuts import render
from .models import Account,Billing
from .forms import AccountsForm,BillingForm
from django.contrib.auth.decorators import login_required
from dms_main.decorators import check_user_client_redirect
from django.contrib import messages
from django.http import HttpResponseRedirect,HttpResponse
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.worksheet.header_footer import HeaderFooter
from openpyxl.styles import Font
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
)

from openpyxl.worksheet.table import Table, TableStyleInfo

@check_user_client_redirect
@login_required(login_url="/")
def account_view(request):

    account_info = Account.objects.all()

    context = {
        "accounts_info":account_info,
    }
    return render(request,"accounts.html",context)

@check_user_client_redirect
@login_required(login_url="/")
def add_account_view(request):

    account_form = AccountsForm()
    
    if request.method == "POST":

        account_form = AccountsForm(request.POST, request.FILES)

        if account_form.is_valid():
            account_form.save()
            messages.success(request,"You have successfully added a new information.") 
            return HttpResponseRedirect("/dms/account/")
        else:
            messages.error(request,"Having an error in adding a new payment, please try again.") 
            return HttpResponseRedirect("/dms/account/")
    
    context = {
        "account_form":account_form,
    }

    return render(request,"add_account.html",context)

@check_user_client_redirect
@login_required(login_url="/")
def edit_account_view(request,id):
    account_instance = Account.objects.get(id=id)
    account_form = AccountsForm(instance=account_instance)

    if request.method == "POST":
        save_account = AccountsForm(request.POST,instance=account_instance)
        
        if save_account.is_valid():
            save_account.save()
            messages.success(request,"You have successfully updated the account information.") 
            return HttpResponseRedirect("/dms/account/")
        else:
            messages.error(request,"Having an error in adding a new payment, please try again.") 
            return HttpResponseRedirect("/dms/account/")
    context = {
        "account_form":account_form
    }
    
    return render(request,"edit_account.html",context)

@check_user_client_redirect
@login_required(login_url="/")
def delete_account_view(request):
    return render(request,"delete_account.html")



@check_user_client_redirect
@login_required(login_url="/")
def export_account(request):
    workbook = openpyxl.Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active
    
    columns = [
        ("A", "FIRSTNAME"),
        ("B", "LASTNAME"),
        ("C", "REFFERENCE CODE"),
        ("D", "DUE DATE"),
        ("E", "PAYMENT AS MONTH OF"),
        ("F", "AMOUNT PAID"),
        ("G", "PAYMENT TYPE"),
        ("H", "PAYMENT STATUS"),
        ("I", "DATE PAID")
    ]

    for col,header in columns:
        cell = worksheet["{}1".format(col)]
        cell.value = header
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True,color="FFFFFF")
        column_letter = get_column_letter(cell.column)
        max_length    = max(len(str(cell.value) if cell.value is not None else 0) for cell in worksheet[column_letter])
        adjusted_width = (max_length + 4)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    row_num = 2


    for account in Account.objects.all():
        worksheet.cell(row=row_num, column=1, value=account.user_id.first_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=2, value=account.user_id.last_name).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=account.ref_code if account.ref_code else 'N/A').alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value=account.billing_id.due_date.strftime('%d/%m/%Y')).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=5, value=account.billing_id.month.strftime('%B')).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=6, value=account.amount_paid).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=7, value=account.payment_type).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=8, value=account.payment_status).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=9, value=account.date_paid.strftime('%d/%m/%Y')).alignment = Alignment(horizontal='center')

        row_num += 1

    table = Table(displayName="PaymentTable", ref=worksheet.dimensions)
    style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    worksheet.add_table(table)

    filename_str = f"payment_informations_{datetime.now()}.xlsx"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'

    workbook.save(response)

    return response


@check_user_client_redirect
@login_required(login_url="/")
def billing_view(request):
    context = {}
    context['billing_list'] = Billing.objects.all().order_by("month")

    print(context['billing_list'])
    return render(request,"billing/billing.html",context)


@check_user_client_redirect
@login_required(login_url="/")
def add_billing_view(request):
    context = {}
    if request.method == "POST":

        create_billing = BillingForm(request.POST)
        
        if create_billing.is_valid():
            create_billing.save()
            print("Billing Saved")
        else:
            print(create_billing.errors)
    context['form'] =  BillingForm()

    return render(request,"billing/add_billing.html",context)


@check_user_client_redirect
@login_required(login_url="/")
def edit_billing_view(request,id):
    instance = Billing.objects.get(id=id)
    context = {}
    
    if request.method == "POST":
        edit_form = BillingForm(request.POST,instance=instance)

        if edit_form.is_valid():
            edit_form.save()
            print("Saved billing")

    context['form'] =  BillingForm(instance=instance)

    return render(request,"billing/edit_billing.html",context)
